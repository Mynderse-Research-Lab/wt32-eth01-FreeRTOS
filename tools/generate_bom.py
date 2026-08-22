"""Generate driver_datasheets_and_calculations/BOM.xlsx (multi-unit standard BOM)."""
from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = (
    Path(__file__).resolve().parents[1]
    / "driver_datasheets_and_calculations"
    / "BOM.xlsx"
)

# Standard BOM columns (IEC / industrial panel shop convention)
BOM_HEADERS: list[tuple[str, int]] = [
    ("Line", 6),
    ("Level", 6),
    ("Category", 16),
    ("Qty", 8),
    ("UOM", 6),
    ("Manufacturer", 18),
    ("Part number", 32),
    ("Description", 40),
    ("Ref des", 10),
    ("Location", 14),
    ("Status", 14),
    ("Source / PO", 18),
    ("Notes", 44),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
SECTION_FILL = PatternFill("solid", fgColor="D9E1F2")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Side(style="thin", color="B4B4B4")

STATUS_FILLS = {
    "On order": PatternFill("solid", fgColor="C6EFCE"),
    "On hand": PatternFill("solid", fgColor="B4C6E7"),
    "Panel design": PatternFill("solid", fgColor="FFEB9C"),
    "Machine-mounted": PatternFill("solid", fgColor="E2EFDA"),
    "Field install": PatternFill("solid", fgColor="BDD7EE"),
    "Excluded": PatternFill("solid", fgColor="F2F2F2"),
}

# Each row: (level, category, qty, uom, mfr, part_no, description, ref_des, location, status, source_po, notes)
# level 0 = section header (category only); level 1 = line item
BOM_ROWS: list[tuple] = [
    # --- 1.0 Enclosure & incoming power ---
    (0, "1.0 Enclosure & incoming", "", "", "", "", "Section — enclosure and mains entry", "", "", "", "", ""),
    (
        1, "Enclosure", 1, "EA", "TBD", "TBD",
        "Control panel enclosure, NEMA / IP rated",
        "ENC1", "Panel", "Panel design", "—",
        "Size for 3 drives, filters, contactors, PSU, terminal rails.",
    ),
    (
        1, "Enclosure", 1, "EA", "TBD", "TBD",
        "Main disconnect, lockable handle",
        "DISC1", "Panel", "Panel design", "—",
        "Upstream of branch breakers; size per utility feed.",
    ),
    # --- 2.0 Power distribution & protection ---
    (0, "2.0 Power distribution & protection", "", "", "", "", "Section — main breaker, distribution, branch protection", "", "", "", "", ""),
    (
        1, "Protection", 1, "EA", "Allen-Bradley", "140UT-D7D3-C40",
        "Main circuit breaker, MCCB, 40 A, 3-pole; applied at 240V (device rated to 480V)",
        "CB0", "Panel", "Panel design", "140UT-TD001",
        "Panel AC input. SUPPLY IS 240V-CLASS 3-PHASE (200-240V) to match Kinetix 5100 E10xx (170-253V) -- NOT 480Y/277V; the 480Y/277V figure is the 1489/140U Type-E SCCR wye rating, not the drive supply. Calc feeder load 20.1 A (NEC 430.24: 1.25x largest INPUT FLC 11.6 A + 2.84 + 1.5 + 1.2 + 0.1). 430.62: main OCPD <= 40 A branch + 5.6 A others = 45.6 A, so 40 A OK (30 A also compliant). For 1489 Type-E self-protection, source must be wye (e.g., 208Y/120); on a delta source verify SCCR separately.",
    ),
    (
        1, "Distribution", 1, "EA", "Allen-Bradley", "1492-PD3141",
        "AC power distribution block, 3-pole, 175 A",
        "PDB1", "Panel", "Panel design", "1492-BR018",
        "AC distribution block feeding the four branch breakers from CB0.",
    ),
    (
        1, "Protection", 1, "EA", "Allen-Bradley", "1489-M2D400",
        "Miniature circuit breaker, 40 A, 3-pole",
        "CB1", "Panel", "On order", "McMc P0078269 L9",
        "X axis AC branch — 2198-E1020-ERS.",
    ),
    (
        1, "Protection", 1, "EA", "Allen-Bradley", "1489-M2D100",
        "Miniature circuit breaker, 10 A, 3-pole",
        "CB2", "Panel", "On order", "McMc P0078269 L18",
        "Z axis AC branch — 2198-E1004-ERS.",
    ),
    (
        1, "Protection", 1, "EA", "Allen-Bradley", "1489-M2D020",
        "Miniature circuit breaker, 2 A, 3-pole",
        "CB3", "Panel", "On order", "McMc P0078269 L8",
        "Theta / HCS01 AC branch. Input FLC ~1.5 A; 2 A may nuisance-trip on inrush -- per NEC 430.52 SCGF may be raised to allow starting (up to ~6 A), e.g. 1489-M2D030/040.",
    ),
    (
        1, "Protection", 1, "EA", "Allen-Bradley", "1489-M2D010",
        "Miniature circuit breaker, 1 A, 3-pole",
        "CB4", "Panel", "On order", "McMc P0078269 L17",
        "AC feed to 24 VDC PSU. NOTE: 1 A is UNDERSIZED -- 1606-XLE240E draws ~1.2 A @240V (~2.3 A @120V) plus SMPS inrush; use >=2 A (1489-M2D020), or 3 A if fed at 120V.",
    ),
    # --- 3.0 Safety switching ---
    (0, "3.0 Safety switching", "", "", "", "", "Section — E-stop controlled AC contactors", "", "", "", "", ""),
    (
        1, "Safety", 1, "EA", "Allen-Bradley", "100S-C43EJ14C",
        "Safety contactor, 3-pole, 24 VDC coil",
        "K1", "Panel", "Panel design", "Allen-Bradley catalog",
        "X mains; between CB1 and DB127-F; E-stop controlled.",
    ),
    (
        1, "Safety", 1, "EA", "Allen-Bradley", "100S-C16EJ14C",
        "Safety contactor, 3-pole, 24 VDC coil",
        "K2", "Panel", "Panel design", "Allen-Bradley catalog",
        "Z mains; between CB2 and DB111-F.",
    ),
    (
        1, "Safety", 1, "EA", "Allen-Bradley", "100S-C09EJ14C",
        "Safety contactor, 3-pole, 24 VDC coil",
        "K3", "Panel", "Panel design", "Allen-Bradley catalog",
        "Theta mains; between CB3 and HCS01.",
    ),
    (
        1, "Safety", 1, "EA", "Allen-Bradley", "800F-X01",
        "E-stop push button, 22.5 mm, 1 NC contact",
        "ES1", "Machine", "Panel design", "Allen-Bradley catalog",
        "Machine-mounted; NC in 24 V contactor coil chain.",
    ),
    (
        1, "Safety", 1, "EA", "Allen-Bradley", "800F-X01",
        "E-stop push button, 22.5 mm, 1 NC contact",
        "ES2", "Machine", "Panel design", "Allen-Bradley catalog",
        "Second station or dual-channel per layout.",
    ),
    # --- 4.0 Line filters ---
    (0, "4.0 Line filters", "", "", "", "", "Section — Kinetix mains filters", "", "", "", "", ""),
    (
        1, "Filter", 1, "EA", "Allen-Bradley", "2198-DB127-F",
        "Kinetix 5100 line filter",
        "LF1", "Panel", "On order", "McMc P0078269 L4",
        "X axis; after K1, before E1020.",
    ),
    (
        1, "Filter", 1, "EA", "Allen-Bradley", "2198-DB111-F",
        "Kinetix 5100 line filter",
        "LF2", "Panel", "On order", "McMc P0078269 L16",
        "Z axis; after K2, before E1004.",
    ),
    (
        1, "Filter", 1, "EA", "Bosch Rexroth", "NFD03.1-480-016",
        "Mains EMC filter, 480 V, 16 A",
        "LF3", "Panel", "Panel design", "Rexroth HCS01 project manual",
        "Theta; after K3, before HCS01. Confirm rating against HCS01 dimensioning. Not on either PO.",
    ),
    # --- 5.0 Servo drives ---
    (0, "5.0 Servo drives", "", "", "", "", "Section — motion drives", "", "", "", "", ""),
    (
        1, "Drive", 1, "EA", "Allen-Bradley", "2198-E1020-ERS",
        "Kinetix 5100 servo drive, PTI, Ethernet",
        "DRV1", "Panel", "On order", "McMc P0078269",
        "X axis — MPL-A320P / Beta 100-ZRS.",
    ),
    (
        1, "Drive", 1, "EA", "Allen-Bradley", "2198-E1004-ERS",
        "Kinetix 5100 servo drive, PTI, Ethernet",
        "DRV2", "Panel", "On order", "McMc P0078269",
        "Z axis — MPL-A310F / Beta 80-SRS.",
    ),
    (
        1, "Drive", 1, "EA", "Bosch Rexroth", "HCS01.1E-W0005-A-03-B-ET-EC-NN-NN",
        "IndraDrive Cs servo drive, Ethernet",
        "DRV3", "Panel", "On order", "Youngblood P077662",
        "Theta — ERD 04-40-D-H-N.",
    ),
    (
        1, "Drive I/O", 1, "EA", "Allen-Bradley", "2198-TBIO",
        "Kinetix 5100 I/O terminal block",
        "TB1", "Panel", "On order", "McMc P0078269 L2",
        "E1020 field signals — 16–30 AWG.",
    ),
    (
        1, "Drive I/O", 1, "EA", "Allen-Bradley", "2198-TBIO",
        "Kinetix 5100 I/O terminal block",
        "TB2", "Panel", "On order", "McMc P0078269 L12",
        "E1004 field signals.",
    ),
    (
        1, "Drive I/O", 1, "EA", "Allen-Bradley", "2198-K51CK-D15M",
        "15-pin MFB feedback connector kit",
        "CON1", "Panel", "On order", "McMc P0078269 L3",
        "E1020 feedback cable termination.",
    ),
    (
        1, "Drive I/O", 1, "EA", "Allen-Bradley", "2198-K51CK-D15M",
        "15-pin MFB feedback connector kit",
        "CON2", "Panel", "On order", "McMc P0078269 L11",
        "E1004 feedback cable termination.",
    ),
    # --- 6.0 Control power ---
    (0, "6.0 Control power", "", "", "", "", "Section — 24 VDC control supply", "", "", "", "", ""),
    (
        1, "Power supply", 1, "EA", "Allen-Bradley", "1606-XLE240E",
        "Switched-mode PSU, 24 VDC, 240 W class",
        "PS1", "Panel", "Panel design", "Rockwell / distributor",
        "≥3 A DC for two Kinetix CP + relay + E-stop.",
    ),
    (
        1, "Control", 1, "EA", "Phoenix Contact", "2900299",
        "PLC interface relay, 24 VDC, 1 CO",
        "CR1", "Panel", "Panel design", "Phoenix Contact",
        "Gripper valve coil driver; GPIO4 cannot sink coil current — use CR1.",
    ),
    (
        1, "Logic power", 1, "EA", "Phoenix Contact", "STEP3-PS/1AC/5DC/3/PT (1170954)",
        "PSU, 1-phase 240 VAC in / 5 VDC 3 A out, DIN/panel mount",
        "PS2", "Panel", "Panel design", "Phoenix Contact",
        "240 VAC -> 5 VDC for Gantry Control Unit. Alt compact: STEP-PS/1AC/5DC/2 (2320513), 5 V 2 A.",
    ),
    (
        1, "Logic power", 1, "EA", "(no Phoenix SKU)", "5 VDC -> 3.3 VDC POL",
        "5 V to 3.3 V regulator — see note",
        "PS3", "Controller box", "Excluded", "Design note",
        "Recommended: power WT32-ETH01 from 5 V rail and use its on-board 3.3 V LDO; no separate panel 3.3 V supply. If a dedicated 3.3 V rail is required, use a board-mount POL (e.g., Traco TSR-1 / Recom), not Phoenix.",
    ),
    # --- 7.0 Terminal blocks (1492-J) ---
    (0, "7.0 Terminal blocks", "", "", "", "", "Section — Allen-Bradley 1492-J distribution (1492-TD015)", "", "", "", "", ""),
    (
        1, "Terminal", 2, "EA", "Allen-Bradley", "1492-J6-BR",
        "Feed-through terminal, 6 mm², brown",
        "XT1", "Panel", "Panel design", "1492-TD015",
        "X branch L1 — #8 AWG; 2 per phase in/out.",
    ),
    (
        1, "Terminal", 2, "EA", "Allen-Bradley", "1492-J6-BL",
        "Feed-through terminal, 6 mm², black",
        "XT2", "Panel", "Panel design", "1492-TD015",
        "X branch L2.",
    ),
    (
        1, "Terminal", 2, "EA", "Allen-Bradley", "1492-J6-OR",
        "Feed-through terminal, 6 mm², orange",
        "XT3", "Panel", "Panel design", "1492-TD015",
        "X branch L3.",
    ),
    (
        1, "Terminal", 3, "EA", "Allen-Bradley", "1492-J6-G",
        "Feed-through terminal, 6 mm², green",
        "XT4", "Panel", "Panel design", "1492-TD015",
        "X branch PE — #8 AWG.",
    ),
    (
        1, "Terminal", 6, "EA", "Allen-Bradley", "1492-J3-BR",
        "Feed-through terminal, 2.5 mm², brown",
        "XT5", "Panel", "Panel design", "1492-TD015",
        "Z / Theta / PSU AC L1 — #16 AWG.",
    ),
    (
        1, "Terminal", 6, "EA", "Allen-Bradley", "1492-J3-BL",
        "Feed-through terminal, 2.5 mm², black",
        "XT6", "Panel", "Panel design", "1492-TD015",
        "Z / Theta / PSU AC L2.",
    ),
    (
        1, "Terminal", 6, "EA", "Allen-Bradley", "1492-J3-OR",
        "Feed-through terminal, 2.5 mm², orange",
        "XT7", "Panel", "Panel design", "1492-TD015",
        "Z / Theta / PSU AC L3.",
    ),
    (
        1, "Terminal", 8, "EA", "Allen-Bradley", "1492-J3-G",
        "Feed-through terminal, 2.5 mm², green",
        "XT8", "Panel", "Panel design", "1492-TD015",
        "Branch PE — #16 AWG.",
    ),
    (
        1, "Terminal", 14, "EA", "Allen-Bradley", "1492-J3-RE",
        "Feed-through terminal, 2.5 mm², red",
        "XT9", "Panel", "Panel design", "1492-TD015",
        "+24 VDC distribution — land #18 AWG.",
    ),
    (
        1, "Terminal", 14, "EA", "Allen-Bradley", "1492-J3-B",
        "Feed-through terminal, 2.5 mm², blue",
        "XT10", "Panel", "Panel design", "1492-TD015",
        "0 VDC / DC common return.",
    ),
    (
        1, "Terminal", 6, "EA", "Allen-Bradley", "1492-J3-Y",
        "Feed-through terminal, 2.5 mm², yellow",
        "XT11", "Panel", "Panel design", "1492-TD015",
        "E-stop NC chain.",
    ),
    (
        1, "Terminal", 4, "EA", "Allen-Bradley", "1492-J3-W",
        "Feed-through terminal, 2.5 mm², white",
        "XT12", "Panel", "Panel design", "1492-TD015",
        "Relay coil / spare DC taps.",
    ),
    (
        1, "Terminal", 12, "EA", "Allen-Bradley", "1492-J3",
        "Feed-through terminal, 2.5 mm², grey",
        "XT13", "Panel", "Panel design", "1492-TD015",
        "Optional field-signal marshalling (#22 AWG).",
    ),
    (
        1, "Terminal acc.", 2, "EA", "Allen-Bradley", "1492-DR5",
        "DIN rail, 1 m, aluminum symmetrical",
        "—", "Panel", "Panel design", "1492-TD015",
        "Terminal block mounting.",
    ),
    (
        1, "Terminal acc.", 4, "SET", "Allen-Bradley", "1492-EAHJ35",
        "End anchor, heavy duty, for J3/J6",
        "—", "Panel", "Panel design", "1492-TD015",
        "One set per rail end.",
    ),
    (
        1, "Terminal acc.", 8, "EA", "Allen-Bradley", "1492-EBJ3",
        "End barrier, grey",
        "—", "Panel", "Panel design", "1492-TD015",
        "Segment AC / DC / PE rails.",
    ),
    (
        1, "Terminal acc.", 6, "EA", "Allen-Bradley", "1492-CJJ5-10",
        "Center jumper, 10-pole, J3",
        "—", "Panel", "Panel design", "1492-TD015",
        "Bus J3 phase and DC strips.",
    ),
    (
        1, "Terminal acc.", 2, "EA", "Allen-Bradley", "1492-CJJ8-10",
        "Center jumper, 10-pole, J6",
        "—", "Panel", "Panel design", "1492-TD015",
        "Bus J6 X-phase strips.",
    ),
    (
        1, "Terminal acc.", 2, "EA", "Allen-Bradley", "1492-M5X12",
        "Marker card, J3, 144 markers",
        "—", "Panel", "Panel design", "1492-TD015",
        "Wire label every block.",
    ),
    (
        1, "Terminal acc.", 1, "EA", "Allen-Bradley", "1492-MR8X12",
        "Marker card, J6, 84 markers",
        "—", "Panel", "Panel design", "1492-TD015",
        "J6 rail labels.",
    ),
    # --- 8.0 Pneumatic circuit ---
    (0, "8.0 Pneumatic circuit", "", "", "", "", "Section — gripper air (KGG 100-80), 6 mm, 6 bar", "", "", "", "", ""),
    (
        1, "Pneumatic", 1, "EA", "Festo", "MS4-LFR-1/4-D6-CR-M-AS",
        "Filter regulator, G1/4, 0.3-7 bar, 5 um, lockable",
        "FRL1", "Panel or machine", "Panel design", "Festo 529144",
        "Air prep; set 6.0 bar flowing. ISO 8573-1 [7:4:4].",
    ),
    (
        1, "Pneumatic", 1, "EA", "Festo", "QS-1/4-6 (153003)",
        "Push-in fitting, G1/4 male to 6 mm tube",
        "PF1", "Panel or machine", "Panel design", "Festo 153003",
        "FRL outlet to supply tube.",
    ),
    (
        1, "Pneumatic", 3, "M", "Festo", "PUN-H-6X1-BL (197384)",
        "Pneumatic tubing, 6 mm OD / 4 mm ID, PU, blue",
        "TB-A1", "Machine", "Panel design", "Festo 197384",
        "FRL -> valve supply. Keep runs short for high-speed actuation.",
    ),
    (
        1, "Pneumatic", 1, "EA", "Festo", "VUVG-L10-M52-AD-M7-1R8L",
        "5/2 solenoid valve, monostable, 24 VDC, M7, spring return, LED+suppressor",
        "SV1", "End effector", "Panel design", "Festo",
        "Energize = close, spring = open. ~330 L/min, ~14 ms. Mount at gripper.",
    ),
    (
        1, "Pneumatic", 3, "EA", "Festo", "QSM-M7-6-I (153321)",
        "Mini push-in fitting, M7 male to 6 mm tube",
        "PF2", "End effector", "Panel design", "Festo 153321",
        "Valve port 1 (supply) + ports 2/4 (work) = 3 ea.",
    ),
    (
        1, "Pneumatic", 2, "EA", "Festo", "U-M7 / AMTE-M-LH-M7",
        "Sintered exhaust silencer, M7",
        "SIL1", "End effector", "Panel design", "Festo",
        "Valve exhaust ports 3/5. Confirm exact silencer vs valve exhaust thread.",
    ),
    (
        1, "Pneumatic", 1, "M", "Festo", "PUN-H-6X1-BL (197384)",
        "Pneumatic tubing, 6 mm OD / 4 mm ID, PU, blue",
        "TB-A2", "End effector", "Panel design", "Festo 197384",
        "Valve work ports -> gripper. <=0.3 m each for speed (or use M3 hose-free).",
    ),
    (
        1, "Pneumatic", 2, "EA", "Festo", "QSM-M5-6 (153306)",
        "Mini push-in fitting, M5 male to 6 mm tube",
        "PF3", "End effector", "Panel design", "Festo 153306",
        "Gripper air ports. CONFIRM KGG 100-80 port thread (M5 vs M3) from SCHUNK drawing.",
    ),
    (
        1, "End effector", 1, "EA", "SCHUNK", "KGG 100-80",
        "Pneumatic parallel gripper, 40 mm stroke/jaw, 45 cm3/dbl stroke",
        "EE1", "End effector", "On order", "Youngblood P077662 (303-066)",
        "2.5-8 bar, nom 6 bar; close 0.15 s / open 0.19 s (jaws only).",
    ),
    (
        1, "Pneumatic (opt)", 2, "EA", "Festo", "GRLA-M5-QS-6-D (193139)",
        "One-way flow control valve, M5, 6 mm (meter-out)",
        "FC1", "End effector", "Panel design", "Festo 193139",
        "Optional jaw-speed tuning. Omit if datasheet times are acceptable.",
    ),
    (
        1, "Pneumatic (opt)", 1, "EA", "SCHUNK", "SDV-P 04",
        "Pressure-maintenance / grip-hold valve",
        "PMV1", "End effector", "Panel design", "SCHUNK",
        "Optional: holds grip on air loss. Adds slight restriction; omit for max-speed bring-up.",
    ),
    (
        1, "Pneumatic (opt)", 1, "EA", "—", "M3 hose-free mounting plate (custom)",
        "Gripper adapter plate with M3 air passages",
        "MP1", "End effector", "Panel design", "Custom / SCHUNK drawing",
        "Preferred high-speed path: <=50 mm air path, replaces TB-A2/PF3.",
    ),
    # --- 9.0 Controller (machine interface) ---
    (0, "9.0 Controller interface", "", "", "", "", "Section — WT32 motion controller (EIP; not in drive cabinet)", "", "", "", "", ""),
    (
        1, "Controller", 1, "EA", "LilyGO / Espressif", "WT32-ETH01",
        "ESP32 Ethernet controller module",
        "PLC1", "Controller box", "Field install", "Project hardware",
        "Gantry firmware; MQTT pick scheduling over LAN8720.",
    ),
    (
        1, "Controller", 1, "EA", "WIZnet", "WIZ850io (W5500)",
        "SPI Ethernet module for EtherNet/IP daisy-chain",
        "ETH1", "Controller box", "Field install", "Project hardware",
        "EIP originator; SPI pins MOSI12/MISO35/SCLK5/CS15/RST14. See docs/HV_LV_SCHEMATICS.md sheet 06.",
    ),
    # --- 9.1 Field I/O (EIP era — no PTI/opto board) ---
    (
        0, "9.1 Field I/O (EIP)", "", "", "", "",
        "Section — drive-side limits + gripper; see docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
        "", "", "", "", "",
    ),
    (
        1, "Network", 2, "EA", "Generic", "Cat5e patch cable",
        "EtherNet/IP daisy-chain jumpers",
        "W-EIP", "Panel / Machine", "Field install", "—",
        "W5500 -> X PORT1; X PORT2 -> Z PORT1; Z PORT2 -> PC uplink (exclusive).",
    ),
    (
        1, "Signal cable", 50, "FT", "Belden", "8723",
        "2-pair 22 AWG individually shielded, 300V",
        "W-LIM", "Machine", "Field install", "Belden 8723",
        "NC limit switches to TBIO INPUT1-4 / HCS01 X31.5-6 (not ESP32).",
    ),
    # --- 10.0 Field wiring (reference lines) ---
    (0, "10.0 Field wiring", "", "", "", "", "Section — cables & harnesses (not panel-mounted)", "", "", "", "", ""),
    (
        1, "Field cable", 1, "EA", "Allen-Bradley", "2090-CPWM7DF-16AF03",
        "Motor power cable, 16 AWG, 3 m, continuous-flex",
        "W1", "Machine", "On order", "McMc P0078269 L6",
        "DRV1 <-> MPL-A320P.",
    ),
    (
        1, "Field cable", 1, "EA", "Allen-Bradley", "2090-CFBM7DF-CEAF03",
        "Motor feedback cable, 3 m",
        "W2", "Machine", "On order", "McMc P0078269 L7",
        "DRV1 <-> MPL encoder.",
    ),
    (
        1, "Field cable", 1, "EA", "Allen-Bradley", "2090-CPWM7DF-16AF03",
        "Motor power cable, 16 AWG, 3 m, continuous-flex",
        "W3", "Machine", "On order", "McMc P0078269 L14",
        "DRV2 <-> MPL-A310F.",
    ),
    (
        1, "Field cable", 1, "EA", "Allen-Bradley", "2090-CFBM7DF-CEAF03",
        "Motor feedback cable, 3 m",
        "W4", "Machine", "On order", "McMc P0078269 L15",
        "DRV2 <-> MPL encoder.",
    ),
    (
        1, "Field cable", 1, "EA", "SCHUNK", "KA GLT1706-LK-00500-1",
        "ERD motor power cable, 5 m",
        "W5", "Machine", "On order", "Youngblood 349-104",
        "DRV3 <-> ERD — factory cable.",
    ),
    (
        1, "Field cable", 1, "EA", "SCHUNK", "KA WWN1208-GK-00500-K02",
        "ERD HIPERFACE encoder cable, 5 m",
        "W6", "Machine", "On order", "Youngblood 349-544",
        "DRV3 <-> ERD feedback.",
    ),
    (
        1, "Panel wire", 60, "FT", "Lapp / LappUSA", "OLFLEX WIRE MS 2.1, 10 mm2 (UL MTW/1015 600V)",
        "8 AWG MTW single conductor (BK/BR/OR + GN-PE)",
        "W-X", "Panel", "Field install", "Lapp 4160-series",
        "X 40 A branch + CB0 feeder. NEC 430.122: conductor >= 125% of drive INPUT FLC = 1.25 x 11.6 = 14.5 A (8 AWG is conservative; NEC min would be 14 AWG). 13.4 A is the drive OUTPUT, not the branch load. Not in inventory.",
    ),
    (
        1, "Panel wire", 120, "FT", "Lapp / LappUSA", "OLFLEX WIRE MS 2.1, 1.5 mm2 (UL MTW/1015 600V)",
        "16 AWG MTW single conductor (BK/BR/OR + GN-PE)",
        "W-AC16", "Panel", "On hand", "Lapp 4160-series",
        "Z + Theta + PSU AC branches. 80%: 16 AWG cap 10 A OCPD -> 8 A cont >= 2.6/2.0/1.2 A loads.",
    ),
    (
        1, "Panel wire", 80, "FT", "Lapp / LappUSA", "OLFLEX WIRE MS 2.1, 1.0 mm2 (UL MTW/1015 600V)",
        "18 AWG MTW single conductor (RD +24V, BU 0V, YE e-stop, WT)",
        "W-DC18", "Panel", "On hand", "Lapp 4160-series",
        "24 VDC bus, E-stop chain, gripper valve coil. 80%: 18 AWG 7 A OCPD -> 5.6 A cont >= ~3 A.",
    ),
]

# Conveyor Control Unit (assembly LT-CONV-CU-001) — stays on the conveyor.
# Replaces the existing potentiometer DC drive with a WT32-controlled drive,
# reads the GHW38 wheel encoder, and publishes belt speed over Ethernet/MQTT.
CONVEYOR_BOM_ROWS: list[tuple] = [
    # --- C1.0 Power & protection ---
    (0, "C1.0 Power & protection", "", "", "", "", "Section — CCU mains entry, supplies, motor-bus protection", "", "", "", "", ""),
    (
        1, "Protection", 1, "EA", "Allen-Bradley", "140U-D6D2-C10",
        "Main circuit breaker / disconnect, MCCB, 10 A, 2-pole",
        "CB-C0", "CCU enclosure", "Panel design", "140U-PP010",
        "CCU 240 VAC input. Total ~4.4 A (24 V PSU ~4.3 A + 5 V PSU ~0.1 A).",
    ),
    (
        1, "Power supply", 1, "EA", "Phoenix Contact", "QUINT4-PS/1AC/24DC/40 (2904602)",
        "PSU, 1-phase 240 VAC in / 24 VDC 40 A (960 W) out",
        "PS-C1", "CCU enclosure", "Panel design", "Phoenix Contact",
        "Motor bus. 80%: motor 20.8 A / 0.8 = 26 A -> 30 A min; 40 A for inrush. REUSE existing 24 V bus if the conveyor already supplies one.",
    ),
    (
        1, "Logic power", 1, "EA", "Phoenix Contact", "STEP3-PS/1AC/5DC/3/PT (1170954)",
        "PSU, 1-phase 240 VAC in / 5 VDC 3 A out, DIN/panel mount",
        "PS-C2", "CCU enclosure", "Panel design", "Phoenix Contact",
        "240 VAC -> 5 VDC for WT32, encoder (5 V), and level shifter.",
    ),
    (
        1, "Protection", 1, "EA", "Bussmann / Allen-Bradley", "30 A fuse + DIN fuse holder",
        "Motor-bus DC fuse, 30 A, with holder",
        "FU-C1", "CCU enclosure", "Panel design", "—",
        "Protects 24 V motor bus / driver. 10 AWG conductor protected at 30 A.",
    ),
    # --- C2.0 Drive & motor ---
    (0, "C2.0 Drive & motor", "", "", "", "", "Section — WT32-controlled brushed DC drive", "", "", "", "", ""),
    (
        1, "Drive", 1, "EA", "Cytron", "MD30C R3",
        "Brushed DC motor driver, 5-30 VDC, 30 A cont / 80 A peak, logic PWM+DIR",
        "DRV-C1", "CCU enclosure", "Panel design", "Cytron",
        "WT32 3.3 V PWM+DIR. 80%: 30 A x 0.8 = 24 A >= 20.8 A load. Alt (industrial DIN, 4Q): Roboteq SBL2360T. Mount on sub-panel standoffs.",
    ),
    (
        1, "Drive acc.", 1, "EA", "—", "Heatsink + 24 VDC fan",
        "Forced-air cooling for driver at ~21 A continuous",
        "HS-C1", "CCU enclosure", "Panel design", "—",
        "Required for continuous 20.8 A through MD30C.",
    ),
    (
        1, "Motor", 1, "EA", "(existing)", "Conveyor brushed DC motor",
        "24 VDC, 500 W, 3000 rpm, 2-wire brushed (~20.8 A FL)",
        "MOT-C1", "Conveyor", "Machine-mounted", "Existing / accounted",
        "Reference only — existing motor retained; generic pot driver replaced by DRV-C1.",
    ),
    # --- C3.0 Controller & feedback ---
    (0, "C3.0 Controller & feedback", "", "", "", "", "Section — WT32 + wheel encoder interface", "", "", "", "", ""),
    (
        1, "Controller", 1, "EA", "LilyGO / Espressif", "WT32-ETH01",
        "ESP32 Ethernet controller module",
        "PLC-C1", "CCU enclosure", "Field install", "Project hardware",
        "Powered from 5 V (PS-C2). PWM+DIR -> DRV-C1; reads encoder; publishes belt speed via MQTT.",
    ),
    (
        1, "Interface", 1, "EA", "Texas Instruments", "TXB0104 (level-shifter module)",
        "4-bit bidirectional 5 V <-> 3.3 V level shifter",
        "LS-C1", "CCU enclosure", "Panel design", "TI / module vendor",
        "Encoder A/B/Z (~5 V push-pull) -> 3.3 V WT32 PCNT. ESP32 is NOT 5 V tolerant. Opto-isolated input is an alternative for noise immunity.",
    ),
    (
        1, "Feedback", 1, "EA", "CALT / Hengxiang", "GHW38-06G500BMP526300",
        "Incremental wheel encoder, 500 PPR, push-pull, 5-26 V, 300 mm wheel",
        "ENC-C1", "Conveyor", "Machine-mounted", "Existing / accounted",
        "Reference only — encoder + mounting accounted. Power at 5 V; A/B/Z via LS-C1.",
    ),
    (
        1, "Network", 1, "EA", "—", "Cat5e patch cable",
        "Ethernet drop, CCU WT32 -> plant switch",
        "W-CNET", "Conveyor", "Field install", "—",
        "Carries belt-speed MQTT to gantry/vision network.",
    ),
    # --- C4.0 Terminal blocks & DIN (1492-J) ---
    (0, "C4.0 Terminal blocks", "", "", "", "", "Section — Allen-Bradley 1492-J distribution", "", "", "", "", ""),
    (
        1, "Terminal", 2, "EA", "Allen-Bradley", "1492-J6-BK",
        "Feed-through terminal, 6 mm², black",
        "XT-C1", "CCU enclosure", "Panel design", "1492-TD015",
        "Motor leads M+/M- (10 AWG). 1492-J6 = #22-8 AWG.",
    ),
    (
        1, "Terminal", 3, "EA", "Allen-Bradley", "1492-J3-RE",
        "Feed-through terminal, 2.5 mm², red",
        "XT-C2", "CCU enclosure", "Panel design", "1492-TD015",
        "+24 V distribution.",
    ),
    (
        1, "Terminal", 3, "EA", "Allen-Bradley", "1492-J3-B",
        "Feed-through terminal, 2.5 mm², blue",
        "XT-C3", "CCU enclosure", "Panel design", "1492-TD015",
        "0 V / DC common.",
    ),
    (
        1, "Terminal", 4, "EA", "Allen-Bradley", "1492-J3",
        "Feed-through terminal, 2.5 mm², grey",
        "XT-C4", "CCU enclosure", "Panel design", "1492-TD015",
        "5 V + encoder/signal marshalling.",
    ),
    (
        1, "Terminal", 3, "EA", "Allen-Bradley", "1492-J3-G",
        "Feed-through terminal, 2.5 mm², green",
        "XT-C5", "CCU enclosure", "Panel design", "1492-TD015",
        "PE / equipment ground.",
    ),
    (
        1, "Terminal acc.", 1, "EA", "Allen-Bradley", "1492-DR5",
        "DIN rail, 1 m, aluminum symmetrical",
        "—", "CCU enclosure", "Panel design", "1492-TD015",
        "Terminal block mounting.",
    ),
    (
        1, "Terminal acc.", 2, "SET", "Allen-Bradley", "1492-EAHJ35",
        "End anchor, heavy duty",
        "—", "CCU enclosure", "Panel design", "1492-TD015",
        "One set per rail end.",
    ),
    (
        1, "Terminal acc.", 1, "EA", "Allen-Bradley", "1492-M5X12",
        "Marker card, J3, 144 markers",
        "—", "CCU enclosure", "Panel design", "1492-TD015",
        "Label every block.",
    ),
    # --- C5.0 Enclosure ---
    (0, "C5.0 Enclosure", "", "", "", "", "Section — CCU enclosure (on conveyor)", "", "", "", "", ""),
    (
        1, "Enclosure", 1, "EA", "TBD", "TBD",
        "Small control enclosure, NEMA 12 / IP rated",
        "ENC-C1", "Conveyor", "Panel design", "—",
        "Size for 2 PSUs, driver+heatsink, WT32, terminal rail. Conveyor-mounted.",
    ),
    (
        1, "Enclosure acc.", 1, "EA", "TBD", "TBD",
        "Wiring duct / wireway + DIN rail kit",
        "—", "Conveyor", "Panel design", "—",
        "Internal wire management.",
    ),
    # --- C6.0 Wiring (80%-sized, part-numbered) ---
    (0, "C6.0 Wiring", "", "", "", "", "Section — CCU conductors (80% criteria)", "", "", "", "", ""),
    (
        1, "Panel wire", 30, "FT", "Lapp / LappUSA", "OLFLEX WIRE MS 2.1, 6 mm2 (UL MTW/1015 600V)",
        "10 AWG MTW single conductor (motor + / -)",
        "W-MOT", "CCU enclosure", "Field install", "Lapp 4160-series",
        "24 V motor leads, 20.8 A. 80%: 10 AWG=35 A @75C, 28 A cont >= 20.8 A. 12 AWG (25 A) too small.",
    ),
    (
        1, "Panel wire", 30, "FT", "Lapp / LappUSA", "OLFLEX WIRE MS 2.1, 1.5 mm2 (UL MTW/1015 600V)",
        "16 AWG MTW single conductor (240 VAC feed + PE)",
        "W-AC16", "CCU enclosure", "On hand", "Lapp 4160-series",
        "240 VAC to PSUs, ~4.4 A. 80%: 16 AWG cap 10 A OCPD -> 8 A cont.",
    ),
    (
        1, "Panel wire", 40, "FT", "Lapp / LappUSA", "OLFLEX WIRE MS 2.1, 1.0 mm2 (UL MTW/1015 600V)",
        "18 AWG MTW single conductor (24 V control + 5 V logic)",
        "W-DC18", "CCU enclosure", "On hand", "Lapp 4160-series",
        "24 V control + 5 V distribution. 80%: 18 AWG 7 A OCPD -> 5.6 A cont.",
    ),
    (
        1, "Signal cable", 1, "EA", "—", "Encoder cable (supplied) / 6-cond shielded",
        "Encoder A/B/Z + Vcc + GND, shielded",
        "W-ENC", "Conveyor", "Field install", "With encoder",
        "GHW38 ships with a cable; extend with 6-conductor shielded (e.g. Belden 9536) if needed.",
    ),
]

BRANCH_HEADERS = [
    ("Circuit", 12),
    ("Breaker", 18),
    ("Wire (Cu)", 14),
    ("Contactor", 18),
    ("Filter", 16),
    ("Load", 32),
    ("Ref des", 10),
    ("Notes", 44),
]

BRANCH_ROWS = [
    ("Panel main", "CB0 / 140UT-D7D3-C40", "8 AWG + PE", "—", "—", "PDB1 distribution", "CB0", "240V 3-ph; feeder load 20.1 A (430.24); 40 A OCPD (<=45.6 A per 430.62)"),
    ("X mains", "CB1 / 40 A", "8 AWG + PE", "K1", "LF1", "2198-E1020-ERS", "DRV1", "Input FLC 11.6 A (240V 3-ph); 8 AWG conservative (430.122 min 14 AWG)"),
    ("Z mains", "CB2 / 10 A", "16 AWG + PE", "K2", "LF2", "2198-E1004-ERS", "DRV2", "Input FLC 2.84 A; 16 AWG (430.122)"),
    ("Theta mains", "CB3 / 2 A", "16 AWG + PE", "K3", "LF3 (NFD03.1)", "HCS01.1E", "DRV3", "Input FLC ~1.5 A; 2 A breaker marginal on inrush -> consider 3-4 A (430.52 allows up to 6 A)"),
    ("24 V PSU", "CB4 / 2 A min", "16 AWG + PE", "—", "—", "1606-XLE240E", "PS1", "Input ~1.2 A @240V (~2.3 A @120V). 1 A breaker UNDERSIZED -> use >=2 A (3 A if 120V tap)"),
    ("Logic PSU", "from CB4 bus", "16 AWG + PE", "—", "—", "STEP3-PS 5 VDC", "PS2", "240 VAC -> 5 VDC for GCU; ~0.1 A, shares CB4 branch"),
    ("24 V DC", "PSU fuse", "18 AWG", "—", "—", "CP, E-stop, CR1", "XT9-XT11", "Not through 1489-M2D"),
    ("CCU 240 VAC in", "CB-C0 / 140U-D6D2-C10", "16 AWG + PE", "—", "—", "24 V + 5 V PSUs", "CB-C0", "Conveyor unit; ~4.4 A total"),
    ("CCU motor bus", "FU-C1 / 30 A DC", "10 AWG", "—", "—", "MD30C -> motor", "DRV-C1", "24 VDC, 20.8 A"),
]

SIZING_HEADERS = [
    ("Circuit", 16),
    ("Load (A)", 10),
    ("Sizing rule", 30),
    ("OCPD", 22),
    ("Conductor", 16),
    ("Ampacity (UL508A T28.1, 60C term)", 18),
    ("80% cont. cap", 14),
    ("Result", 28),
]

SIZING_ROWS = [
    ("Panel main (CB0)", "20.1", "430.24: 1.25x largest INPUT FLC + sum others", "140UT-D7D3-C40 (40 A)", "8 AWG", "40 A", "32 A", "240V 3-ph; 8 AWG=40A(60C)>=20.1A; OCPD<=45.6A (430.62); 8 AWG conservative"),
    ("X mains (E1020)", "11.6", "430.122 cond>=125% input FLC; 430.130/.52 SCGF", "1489-M2D400 (40 A)", "8 AWG", "40 A", "32 A", "40 A <= 400%xFLC=46.4A OK; cond min 14 AWG, 8 AWG conservative"),
    ("Z mains (E1004)", "2.84", "430.122 cond>=125% input FLC", "1489-M2D100 (10 A)", "16 AWG", "10 A", "8 A", "10 A <= 400%=11.4A OK; 16 AWG OK"),
    ("Theta mains (HCS01)", "1.5", "430.122/.130; SCGF must permit start", "1489-M2D020 (2 A) -> 3-4 A", "16 AWG", "10 A", "8 A", "2 A may nuisance-trip inrush; 430.52 allows up to 6 A -> M2D030/040"),
    ("24 V PSU AC", "1.2", "125% incl. SMPS inrush", "1489-M2D020 (2 A) min", "16 AWG", "10 A", "8 A", "M2D010 (1 A) UNDERSIZED < 1.2 A load; use >=2 A (3 A if 120V tap)"),
    ("24 V DC bus", "3.0", "Load <= 0.8 x UL508A ampacity; DC fuse", "PSU DC fuse (~5 A)", "18 AWG", "7 A", "5.6 A", "18 AWG OK"),
    ("Valve coil", "0.3", "Load <= 0.8 x UL508A ampacity", "shared DC branch", "18 AWG", "7 A", "5.6 A", "18 AWG OK"),
    ("Signal", "<0.1", "Signal-level, shielded", "n/a", "22 AWG STP", "—", "—", "Belden 8723"),
    ("CCU motor 24 V", "20.8", "Load <= 0.8 x UL508A ampacity; DC fuse", "30 A DC fuse", "10 AWG", "30 A", "24 A", "10 AWG=30A(60C), 80%=24A>=20.8A; 12 AWG=25A too small"),
    ("CCU 24 V PSU AC", "4.3", "Load <= 0.8 x UL508A ampacity", "140U-D6D2-C10 (10 A)", "16 AWG", "10 A", "8 A", "16 AWG OK"),
    ("CCU 5 V logic", "<1", "Load <= 0.8 x UL508A ampacity", "PSU internal", "18 AWG", "7 A", "5.6 A", "18 AWG OK"),
]

# UL 508A Supplement SB - Short-Circuit Current Rating (SCCR).
# Panel SCCR = lowest-rated branch in the worst current path (after any
# combination ratings). Available fault current at the panel must be <= this.
SCCR_HEADERS = [
    ("Branch / component", 26),
    ("Device", 28),
    ("SCCR (rms sym)", 16),
    ("Basis / note", 56),
]

SCCR_ROWS = [
    ("X / Z drives", "Kinetix 5100 2198-E1020/E1004-ERS", "5 kA", "Datasheet KNX-TD003: 5,000 A rms sym @240V. THIS is the limiting component -> panel SCCR caps at 5 kA unless raised by a listed combination."),
    ("Theta drive", "Rexroth HCS01.1E-W0005", "42 kA", "Datasheet: SCCR 42,000 A. Not limiting."),
    ("Branch protection", "1489-M2D / 140U-D6", "—", "1489/140U as Type-E self-protected combination need a WYE source (e.g., 208Y/120) per 140-TD005; raises combined SCCR but only on wye."),
    ("Main breaker", "140UT-D7D3-C40", "65 kA @240V (typ)", "Confirm interrupting/SCCR vs 140UT-TD001 for the chosen voltage."),
    ("Distribution block", "1492-PD3141", ">=10 kA", "Verify component SCCR contribution per SB4."),
    ("Panel (assembled)", "LT-BPG-PNL-001", "5 kA (default)", "Governed by the 5 kA Kinetix drives. To mark higher, use the drive's listed fuse/breaker combination from KNX-TD003 ahead of each drive."),
    ("Install requirement", "Available fault current", "<= panel SCCR", "Per NEC 409.22 / 110.10 the available fault current at the panel must not exceed the marked SCCR. Obtain utility/transformer let-through; add note to panel nameplate."),
]

REV_HEADERS = [
    ("Rev", 6),
    ("Date", 12),
    ("Author", 16),
    ("Description", 48),
]

REV_ROWS = [
    ("A", str(date.today()), "wt32-eth01-base", "Initial standard-format panel BOM"),
    ("B", str(date.today()), "wt32-eth01-base", "Added Conveyor Control Unit; split into per-unit sheets; renamed to BOM.xlsx"),
]


def _border_range(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            ws.cell(row=r, column=c).border = Border(
                left=THIN, right=THIN, top=THIN, bottom=THIN
            )


def build_document_control(ws) -> None:
    ws.title = "Document control"
    ws["A1"] = "Lawrence Tech Battery Pick-and-Place Line"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Multi-unit Bill of Materials"
    ws["A2"].font = SECTION_FONT
    meta = [
        ("Project", "Lawrence Tech battery pick-and-place line"),
        ("Firmware repo", "wt32-eth01-base"),
        ("Gantry panel", "LT-BPG-PNL-001 — 3-axis gantry control panel + machine interfaces"),
        ("Conveyor panel", "LT-CONV-CU-001 — conveyor control unit (WT32 drive + wheel encoder), on conveyor"),
        ("Vision system panel", "Existing — in place; parts accounted externally (stub sheet)"),
        ("Format", "Standard multi-level BOM (Line / Level / Category / Qty / UOM / Mfr / PN)"),
        ("Primary sheets", "Gantry panel, Conveyor panel, Vision system panel"),
    ]
    for i, (k, v) in enumerate(meta, start=4):
        ws.cell(row=i, column=1, value=k).font = SECTION_FONT
        ws.cell(row=i, column=2, value=v).alignment = WRAP
        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=5)

    rev_start = 4 + len(meta) + 2
    ws.cell(row=rev_start, column=1, value="Revision history").font = SECTION_FONT
    for col, (title, width) in enumerate(REV_HEADERS, start=1):
        cell = ws.cell(row=rev_start + 1, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = width
    for i, row in enumerate(REV_ROWS, start=rev_start + 2):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val).alignment = WRAP

    legend = rev_start + 2 + len(REV_ROWS) + 2
    ws.cell(row=legend, column=1, value="Status legend").font = SECTION_FONT
    for j, (status, meaning) in enumerate(STATUS_FILLS.items(), start=legend + 1):
        ws.cell(row=j, column=1, value=status)
        ws.cell(row=j, column=1).fill = meaning
        ws.cell(row=j, column=2, value={
            "On order": "On McMc or Youngblood PO",
            "On hand": "Project inventory",
            "Panel design": "Defined; not on PO",
            "Machine-mounted": "On machine (gantry/conveyor); existing or wired to panel",
            "Field install": "Field-built harness or wire",
            "Excluded": "Explicitly out of scope",
        }.get(status, ""))
        ws.merge_cells(start_row=j, start_column=2, end_row=j, end_column=5)

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 48


def build_unit_bom(ws, sheet_title: str, title_text: str, rows: list[tuple]) -> None:
    ws.title = sheet_title
    ws["A1"] = title_text
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(BOM_HEADERS))}1")

    header_row = 3
    for col, (title, width) in enumerate(BOM_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = (
        f"A{header_row}:{get_column_letter(len(BOM_HEADERS))}{header_row}"
    )

    line = 0
    data_start = header_row + 1
    for row_idx, row in enumerate(rows, start=data_start):
        level = row[0]
        if level == 0:
            ws.cell(row=row_idx, column=3, value=row[1]).font = SECTION_FONT
            ws.cell(row=row_idx, column=8, value=row[6])
            for c in range(1, len(BOM_HEADERS) + 1):
                ws.cell(row=row_idx, column=c).fill = SECTION_FILL
            continue

        line += 1
        values = [line, *row[:12]]
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = WRAP
            if col == 11:
                fill = STATUS_FILLS.get(str(value))
                if fill:
                    cell.fill = fill

    _border_range(ws, header_row, data_start + len(rows) - 1, 1, len(BOM_HEADERS))


def build_vision_stub(ws) -> None:
    ws.title = "Vision system panel"
    ws["A1"] = "Vision System Control Unit — BOM (reference)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(f"A1:{get_column_letter(len(BOM_HEADERS))}1")
    ws["A2"] = (
        "Vision system control unit is already in place; all parts are accounted for externally. "
        "This sheet is a placeholder — paste vision line items here if they need to be tracked in this workbook."
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells(f"A2:{get_column_letter(len(BOM_HEADERS))}2")

    header_row = 4
    for col, (title, width) in enumerate(BOM_HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    note = ws.cell(row=header_row + 1, column=3, value="Accounted for externally — no line items tracked here.")
    note.font = SECTION_FONT


def build_branch_schedule(ws) -> None:
    ws.title = "Branch schedule"
    ws["A1"] = "Electrical branch circuit schedule (reference)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    start = 3
    for col, (title, width) in enumerate(BRANCH_HEADERS, start=1):
        cell = ws.cell(row=start, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"
    for i, row in enumerate(BRANCH_ROWS, start=start + 1):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val).alignment = WRAP


def build_wiring_reference(ws) -> None:
    ws.title = "Wiring reference"
    ws["A1"] = "Control signal reference (WT32-ETH01 / W5500 EIP)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    headers = [
        ("Signal", 22),
        ("From", 24),
        ("To", 28),
        ("Wire", 14),
        ("Ref / pin", 20),
        ("Notes", 40),
    ]
    rows = [
        ("EIP O→T / T→O", "W5500", "Kinetix X/Z", "Cat5e+", "104 / 154", "Class 1 RPI 5 ms"),
        ("EIP daisy continue", "X PORT2", "Z PORT1", "Cat5e+", "—", "Then Z PORT2 → PC"),
        ("W5500 SPI", "WT32", "WIZ850io", "short harness", "12/35/5/15/14", "MOSI/MISO/SCLK/CS/RST"),
        ("Limits X", "Beta 100-ZRS", "TBIO INPUT1/2", "22 AWG STP", "pins 9/10", "NC sinking"),
        ("Limits Z", "Beta 80-SRS", "TBIO INPUT3/4", "22 AWG STP", "pins 34/8", "NC sinking"),
        ("Gripper", "GPIO4", "CR1 coil → SV1", "18 AWG", "PIN_GRIPPER", "activeHigh = close"),
        ("MQTT Ethernet", "WT32 LAN8720", "Plant network", "Cat5e/6", "—", "Separate from EIP"),
    ]
    start = 3
    for col, (title, width) in enumerate(headers, start=1):
        cell = ws.cell(row=start, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A4"
    for i, row in enumerate(rows, start=start + 1):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val).alignment = WRAP


def build_sizing_basis(ws) -> None:
    ws.title = "Sizing basis"
    ws["A1"] = "Conductor & breaker sizing (UL 508A industrial control panel; NEC for the feeder)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "GOVERNING STANDARD: UL 508A (industrial control panel). Internal-wiring ampacity per UL 508A "
        "Table 28.1 (18 AWG = 7 A, 16 AWG = 10 A); power conductors taken at the 60C termination column "
        "for <=100 A circuits (8 AWG = 40 A, 10 AWG = 30 A) per UL 508A / NEC 110.14(C). NEC governs the "
        "feeder TO the panel, grounding, and field wiring. "
        "SUPPLY: 240V-class 3-phase (200-240V) to match Kinetix 5100 E10xx (170-253V). The earlier "
        "'480Y/277V' was the 1489/140U Type-E SCCR wye rating, NOT the drive supply -- the E10xx drives "
        "cannot take 480V. Use a wye source (e.g., 208Y/120) for 1489 Type-E self-protection. "
        "DRIVES sized on RATED INPUT current, not output (NEC 430.122: conductor >= 125% input FLC; "
        "430.130/430.52: branch SCGF per drive marking, inverse-time CB up to 250%, to 400% to allow start). "
        "Input FLC @230V 3-ph: X (E1020) 11.6 A, Z (E1004) 2.84 A, Theta (HCS01-W0005) ~1.5 A; PSU ~1.2 A. "
        "(Drive OUTPUT 13.4/2.6 A is the motor-side rating and is NOT the branch load.) "
        "Main per 430.24: 1.25 x 11.6 + (2.84+1.5+1.2+0.1) = 20.1 A; main OCPD per 430.62 <= 45.6 A. "
        "Conveyor (CCU): motor 20.8 A (24 VDC 500 W), DC-side fused."
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:H2")
    start = 4
    for col, (title, width) in enumerate(SIZING_HEADERS, start=1):
        cell = ws.cell(row=start, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A5"
    last = start
    for i, row in enumerate(SIZING_ROWS, start=start + 1):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val).alignment = WRAP
        last = i

    # --- Short-circuit current rating (UL 508A Supplement SB) ---
    sccr_title = last + 2
    ws.cell(row=sccr_title, column=1,
            value="Short-circuit current rating (UL 508A Supplement SB)").font = SECTION_FONT
    ws.merge_cells(start_row=sccr_title, start_column=1, end_row=sccr_title, end_column=8)
    sccr_hdr = sccr_title + 1
    for col, (title, _width) in enumerate(SCCR_HEADERS, start=1):
        cell = ws.cell(row=sccr_hdr, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for i, row in enumerate(SCCR_ROWS, start=sccr_hdr + 1):
        for col, val in enumerate(row, start=1):
            ws.cell(row=i, column=col, value=val).alignment = WRAP


def main() -> None:
    wb = Workbook()
    build_document_control(wb.active)
    build_unit_bom(
        wb.create_sheet(), "Gantry panel",
        "LT-BPG-PNL-001 — Gantry Control Panel BOM", BOM_ROWS,
    )
    build_unit_bom(
        wb.create_sheet(), "Conveyor panel",
        "LT-CONV-CU-001 — Conveyor Control Unit BOM", CONVEYOR_BOM_ROWS,
    )
    build_vision_stub(wb.create_sheet())
    build_branch_schedule(wb.create_sheet())
    build_sizing_basis(wb.create_sheet())
    build_wiring_reference(wb.create_sheet())
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
