"""Generate driver_datasheets_and_calculations/WIRE_SIZE_SELECTION.xlsx."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

PROVIDED_FILL = PatternFill("solid", fgColor="C6EFCE")
NOT_PROVIDED_FILL = PatternFill("solid", fgColor="FFC7CE")
PARTIAL_FILL = PatternFill("solid", fgColor="FFEB9C")
INVENTORY_OK_FILL = PatternFill("solid", fgColor="B4C6E7")
INVENTORY_BUY_FILL = PatternFill("solid", fgColor="F4B084")

OUT = (
    Path(__file__).resolve().parents[1]
    / "driver_datasheets_and_calculations"
    / "WIRE_SIZE_SELECTION.xlsx"
)

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")

COLUMNS = [
    ("Item", 6),
    ("Subsystem", 14),
    ("Cable / function", 28),
    ("From", 22),
    ("To", 22),
    ("Wire size", 16),
    ("Catalog / part no.", 32),
    ("Provided on order", 16),
    ("Order line / PO", 24),
    ("Qty", 6),
    ("Length", 10),
    ("Branch / breaker", 18),
    ("Current basis", 24),
    ("Notes", 40),
    ("Source", 36),
]


def set_headers(ws, row: int = 1) -> None:
    for col, (title, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=row, column=col, value=title)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1).coordinate
    ws.auto_filter.ref = f"A{row}:{get_column_letter(len(COLUMNS))}{row}"


def add_row(ws, row: int, values: tuple) -> None:
    provided_col = 8
    for col, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = WRAP
        if col == provided_col:
            fill = {
                "Yes": PROVIDED_FILL,
                "No": NOT_PROVIDED_FILL,
                "Kit only": PARTIAL_FILL,
                "Built-in": PROVIDED_FILL,
            }.get(str(value))
            if fill:
                cell.fill = fill


def build_summary(ws) -> None:
    ws.title = "Summary"
    ws["A1"] = "Gantry wire / cable selection"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    rows = [
        ("Project", "Lawrence Tech battery pick-and-place gantry (wt32-eth01-base)"),
        ("Generated from", "Order PDFs + Allen-Bradley / SCHUNK datasheets in driver_datasheets_and_calculations/"),
        ("", ""),
        ("Axis", "Actuator", "Drive + motor", "Confirmed motor cable AWG"),
        ("X (horizontal belt)", "SCHUNK Beta 100-ZRS", "2198-E1020-ERS + MPL-A320P-SK72AA + Neugart 5:1", "16 AWG"),
        ("Z (vertical ballscrew)", "SCHUNK Beta 80-SRS", "2198-E1004-ERS + MPL-A310F-SK72AA + Neugart 5:1", "16 AWG"),
        ("Theta (rotary)", "SCHUNK ERD 04-40-D-H-N", "HCS01.1E-W0005-A-03-B-ET-EC-NN-NN", "Use SCHUNK factory cable"),
        ("Gripper", "SCHUNK KGG 100-80", "Pneumatic (24 V valve from ESP32 GPIO4)", "6 mm pneumatic / 18-22 AWG valve"),
        ("", "", "", ""),
        ("Key rule", "Motor power AWG follows Rockwell 2090 catalog for the ordered MPL motor, not the smallest gauge the drive terminal accepts."),
        ("Key rule", "Branch/mains wire is sized from the ordered breaker/fuse, not from the drive connector range alone."),
        ("Key rule", "Motion control is EtherNet/IP only — no PTI/opto production harness."),
        ("Key rule", "Limit switches wire to drive TBIO / HCS01 X31 (NC sinking), not to the ESP32."),
        ("Key rule", "See docs/HV_LV_SCHEMATICS.md and docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md."),
        ("", "", "", ""),
        ("Order-provided cabling", "See sheet 'Order-provided cabling' for McMc PO P0078269 and Youngblood OA 1693058 line items."),
        ("", "", "", ""),
        ("On-hand inventory", "Wire: 16 / 18 AWG, 2-pair 22 AWG. Breakers: 1489-M2D010/020/100/400 — see inventory + panel breaker sheets."),
    ]
    for r, row in enumerate(rows, start=3):
        for c, val in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=val).alignment = WRAP
    for col, width in zip("ABCD", [18, 34, 42, 22]):
        ws.column_dimensions[col].width = width


def build_detail_sheet(ws, title: str, rows: list[tuple]) -> None:
    ws.title = title
    set_headers(ws)
    for i, row in enumerate(rows, start=2):
        add_row(ws, i, row)


X_ROWS = [
    (
        1, "X axis", "AC mains (branch)",
        "Panel breaker 1489-M2D400", "2198-E1020-ERS L1/L2/L3",
        "8 AWG Cu min (40 A branch)", "—", "No", "—",
        1, "As required", "40 A breaker on McMc order",
        "E1020 mains; NEC branch sizing",
        "Breaker ordered (McMc L9); branch wire is field-installed.",
        "McMc order acknowledgement.pdf",
    ),
    (
        2, "X axis", "Motor power U/V/W + PE",
        "2198-E1020-ERS motor connector", "MPL-A320P-SK72AA (via Neugart 5:1 to Beta 100-ZRS)",
        "16 AWG", "2090-CPWM7DF-16AF03", "Yes", "McMc L6; PO P0078269",
        1, "3 m", "—", "MPL-A320P 1.3 kW; E1020 Icont 13.4 A rms",
        "Continuous-flex MP-Series cable. Drive end + motor end connectors included.",
        "McMc order acknowledgement.pdf",
    ),
    (
        3, "X axis", "Motor feedback (serial)",
        "2198-E1020-ERS MFB", "MPL-A320P encoder",
        "Signal pairs (not power AWG)", "2090-CFBM7DF-CEAF03", "Yes", "McMc L7; PO P0078269",
        1, "3 m", "—", "Hiperface / MPL feedback",
        "Factory feedback cable. Terminate drive end with 2198-K51CK-D15M (McMc L3).",
        "McMc order acknowledgement.pdf",
    ),
    (
        4, "X axis", "Feedback connector kit (drive end)",
        "2090-CFBM7DF-CEAF03", "2198-E1020-ERS MFB",
        "—", "2198-K51CK-D15M", "Yes", "McMc L3; PO P0078269",
        1, "—", "—", "15-pin feedback termination",
        "Connector kit only — not a field harness. Use with ordered feedback cable.",
        "McMc order acknowledgement.pdf",
    ),
    (
        5, "X axis", "EtherNet/IP (PORT1)",
        "Upstream daisy-chain (W5500 or prior drive PORT2)",
        "2198-E1020-ERS PORT1",
        "Cat5e+", "—", "No", "—",
        1, "As required", "—", "Class 1 EIP",
        "Motion command/feedback via assemblies 104/154. No PTI production wiring.",
        "docs/LOW_LEVEL_GANTRY_CONTROL.md",
    ),
    (
        6, "X axis", "EtherNet/IP (PORT2)",
        "2198-E1020-ERS PORT2",
        "Next hop (Z PORT1 or PC)",
        "Cat5e+", "—", "No", "—",
        1, "As required", "—", "Class 1 EIP",
        "Daisy-chain continue.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
    (
        7, "X axis", "Limit switches MIN/MAX",
        "Beta 100-ZRS NC limits",
        "TBIO INPUT1 pin9 / INPUT2 pin10",
        "22-24 AWG STP", "—", "No", "—",
        2, "As required", "—", "24 V sinking NC",
        "Assign Forward/Reverse Limit in KNX5100C. Not wired to ESP32.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
    (
        8, "X axis", "24 V control power",
        "24 VDC supply", "2198-E1020-ERS CP (+/-)",
        "16-24 AWG", "1606-XLxxx (typical)", "No", "—",
        1, "As required", "—", "~1.4 A @ 24 V per drive family",
        "24 V PSU and wiring not on either order.",
        "AB_Kinetix5100_user_manual.pdf",
    ),
    (
        9, "X axis", "Line filter",
        "Panel / mains", "2198-E1020-ERS",
        "Per filter terminations", "2198-DB127-F", "Yes", "McMc L4; PO P0078269",
        1, "—", "—", "—",
        "Filter hardware on order; panel wiring to filter is field-installed.",
        "McMc order acknowledgement.pdf",
    ),
    (
        10, "X axis", "I/O terminal block",
        "—", "2198-E1020-ERS I/O connector",
        "16-30 AWG at TBIO", "2198-TBIO", "Yes", "McMc L2; PO P0078269",
        1, "—", "—", "—",
        "TBIO for limits / field I/O — not PTI production harness.",
        "McMc order acknowledgement.pdf",
    ),
]

Z_ROWS = [
    (
        1, "Z axis", "AC mains (branch)",
        "Panel breaker 1489-M2D100", "2198-E1004-ERS L1/L2/L3",
        "14 AWG Cu min (10 A branch; Rockwell rec. 15 A)", "—", "No", "—",
        1, "As required", "10 A breaker on McMc order",
        "E1004 mains; fuse table recommends 15 A",
        "Breaker ordered (McMc L18); branch wire field-installed.",
        "McMc order acknowledgement.pdf",
    ),
    (
        2, "Z axis", "Motor power U/V/W + PE",
        "2198-E1004-ERS motor connector", "MPL-A310F-SK72AA (via Neugart 5:1 to Beta 80-SRS)",
        "16 AWG", "2090-CPWM7DF-16AF03", "Yes", "McMc L14; PO P0078269",
        1, "3 m", "—", "MPL-A310F 0.46 kW; E1004 Icont 2.6 A rms",
        "Continuous-flex MP-Series cable.",
        "McMc order acknowledgement.pdf",
    ),
    (
        3, "Z axis", "Motor feedback (serial)",
        "2198-E1004-ERS MFB", "MPL-A310F encoder",
        "Signal pairs", "2090-CFBM7DF-CEAF03", "Yes", "McMc L15; PO P0078269",
        1, "3 m", "—", "MPL feedback",
        "Factory feedback cable. Terminate drive end with 2198-K51CK-D15M (McMc L11).",
        "McMc order acknowledgement.pdf",
    ),
    (
        4, "Z axis", "Feedback connector kit (drive end)",
        "2090-CFBM7DF-CEAF03", "2198-E1004-ERS MFB",
        "—", "2198-K51CK-D15M", "Yes", "McMc L11; PO P0078269",
        1, "—", "—", "15-pin feedback termination",
        "Connector kit only — use with ordered Z feedback cable.",
        "McMc order acknowledgement.pdf",
    ),
    (
        5, "Z axis", "EtherNet/IP (PORT1)",
        "Upstream daisy-chain (X PORT2 or W5500)",
        "2198-E1004-ERS PORT1",
        "Cat5e+", "—", "No", "—",
        1, "As required", "—", "Class 1 EIP",
        "Motion command/feedback via assemblies 104/154. No PTI production wiring.",
        "docs/LOW_LEVEL_GANTRY_CONTROL.md",
    ),
    (
        6, "Z axis", "EtherNet/IP (PORT2)",
        "2198-E1004-ERS PORT2",
        "Next hop (Theta PORT1 or PC)",
        "Cat5e+", "—", "No", "—",
        1, "As required", "—", "Class 1 EIP",
        "Daisy-chain continue.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
    (
        7, "Z axis", "Limit switches MIN/MAX",
        "Beta 80-SRS NC limits",
        "TBIO INPUT3 pin11 / INPUT4 pin12",
        "22-24 AWG STP", "—", "No", "—",
        2, "As required", "—", "24 V sinking NC",
        "Assign Forward/Reverse Limit in KNX5100C. Not wired to ESP32.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
    (
        8, "Z axis", "24 V control power",
        "24 VDC supply", "2198-E1004-ERS CP (+/-)",
        "16-24 AWG", "1606-XLxxx (typical)", "No", "—",
        1, "As required", "—", "~1.4 A @ 24 V per drive family",
        "24 V PSU and wiring not on either order.",
        "AB_Kinetix5100_user_manual.pdf",
    ),
    (
        9, "Z axis", "Line filter",
        "Panel / mains", "2198-E1004-ERS",
        "Per filter terminations", "2198-DB111-F", "Yes", "McMc L16; PO P0078269",
        1, "—", "—", "—",
        "Filter hardware on order; panel wiring field-installed.",
        "McMc order acknowledgement.pdf",
    ),
    (
        10, "Z axis", "I/O terminal block",
        "—", "2198-E1004-ERS I/O connector",
        "16-30 AWG at TBIO", "2198-TBIO", "Yes", "McMc L12; PO P0078269",
        1, "—", "—", "—",
        "TBIO for limits / field I/O — not PTI production harness.",
        "McMc order acknowledgement.pdf",
    ),
]

THETA_ROWS = [
    (
        1, "Theta", "AC mains (branch)",
        "Panel breaker (size per panel design)", "HCS01.1E-W0005-A-03",
        "14 AWG Cu typical (≤15 A branch)", "—", "No", "—",
        1, "As required", "TBD", "HCS01 2 A cont / 5 A peak",
        "Drive on Youngblood order; mains branch wire not included.",
        "OA 1693058",
    ),
    (
        2, "Theta", "Motor power",
        "HCS01 drive", "ERD 04-40-D-H-N",
        "SCHUNK factory cable", "KA GLT1706-LK-00500-1", "Yes", "Youngblood 349-104; PO P077662",
        1, "5 m", "—", "0.43 A nom / 1.29 A max",
        "Factory ERD power cable — do not substitute ad hoc AWG.",
        "OA 1693058 LAWRENCE TECH BATTERY PROJECT 2.9.pdf",
    ),
    (
        3, "Theta", "Encoder (HIPERFACE)",
        "ERD 04-40-D-H-N", "HCS01 feedback",
        "SCHUNK factory cable", "KA WWN1208-GK-00500-K02", "Yes", "Youngblood 349-544; PO P077662",
        1, "5 m", "—", "HIPERFACE multiturn",
        "Factory encoder cable for ERD/HCS01.",
        "OA 1693058 LAWRENCE TECH BATTERY PROJECT 2.9.pdf",
    ),
    (
        4, "Theta", "Cable feed trough (mechanical routing)",
        "—", "ERD 04-40-D-H-N body",
        "Built-in", "—", "Built-in", "Youngblood line 331220",
        1, "—", "—", "—",
        "ERD ordered with pneum. + el. feed trough — routes cables through unit, not a harness.",
        "OA 1693058",
    ),
    (
        5, "Theta", "EtherNet/IP (deferred)",
        "Upstream daisy-chain (Z PORT2)",
        "HCS01.1E multi-Ethernet PORT",
        "Cat5e+", "—", "No", "—",
        1, "As required", "—", "Class 1 EIP (future)",
        "Theta EIP originator path deferred — factory SCHUNK power/encoder cables only for now. No PTI/step-dir production wiring.",
        "docs/LOW_LEVEL_GANTRY_CONTROL.md",
    ),
    (
        6, "Theta", "Limit switches (optional)",
        "ERD / axis NC limits",
        "HCS01 X31.5 / X31.6 digital inputs",
        "22-24 AWG STP", "—", "No", "—",
        2, "As required", "—", "24 V sinking NC",
        "Optional drive-side limits on X31. Not wired to ESP32. Confirm X31 pin map in IndraDrive doc.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
]

GRIPPER_ROWS = [
    (
        1, "Gripper", "Compressed air supply",
        "Plant air / FRL", "KGG 100-80 (M3 direct or port)",
        "6 mm OD pneumatic hose (if not hose-free)", "—", "No", "—",
        2, "As required", "—", "2.5-8 bar (nom 6 bar)",
        "KGG on Youngblood order (303-066) — no hose kit included. Use M3 plate bores or buy 6 mm tube.",
        "Technical Information_SCHUNK KGG 100-80 Parallel Gripper.pdf",
    ),
    (
        2, "Gripper", "Solenoid valve coil",
        "ESP32 GPIO4 -> CR1 (24 V relay) -> valve coil",
        "24 VDC pneumatic solenoid",
        "18-22 AWG", "—", "No", "—",
        1, "As required", "—", "Typical 24 V pneumatic valve",
        "GPIO4 drives CR1 coil side; CR1 contacts switch 24 V to valve. Not MCP23S17. Valve/relay not on either order.",
        "gantry_app_constants.h",
    ),
]

CONTROLLER_ROWS = [
    (
        1, "Controller", "W5500 SPI (short harness)",
        "WT32-ETH01 SPI pins", "W5500 Ethernet controller",
        "26-28 AWG (short harness)", "—", "No", "—",
        4, "<0.3 m", "—", "SPI + reset/CS",
        "Keep SPI run short on controller PCB/harness. No Motion I/O IF / MCP production path.",
        "docs/LOW_LEVEL_GANTRY_CONTROL.md",
    ),
    (
        2, "Controller", "EtherNet/IP originator link",
        "W5500 RJ45", "X drive PORT1 (daisy-chain start)",
        "Cat5e+", "—", "No", "—",
        1, "As required", "—", "Class 1 EIP",
        "W5500 is EIP originator; X->Z->(Theta deferred) daisy-chain. Plant/MQTT may use WT32 RMII separately.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
    (
        3, "Controller", "Gripper GPIO -> relay",
        "WT32-ETH01 GPIO4", "CR1 coil (24 V relay)",
        "22 AWG", "—", "No", "—",
        1, "<0.5 m", "—", "3.3 V logic -> relay coil driver",
        "CR1 contacts feed 24 V to pneumatic valve. See Gripper sheet.",
        "gantry_app_constants.h",
    ),
    (
        4, "Controller", "Plant Ethernet (optional)",
        "WT32-ETH01 RMII PHY", "Plant network / MQTT broker",
        "Cat5e/Cat6 STP", "1585J-M8CBJM-x", "No", "—",
        1, "As required", "—", "—",
        "Standard patch cable — not on order. Separate from W5500 EIP daisy-chain.",
        "pinout.csv / docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
]

IF_BOARD_ROWS = [
    (
        1, "EIP network", "Cat5e daisy-chain",
        "W5500 -> X PORT1 -> X PORT2 -> Z PORT1 -> Z PORT2",
        "(Theta EIP deferred)",
        "Cat5e+", "—", "No", "—",
        3, "As required", "—", "Class 1 EIP",
        "No opto / PTI / MCP Motion I/O IF harness. Motion via EIP assemblies only.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
    (
        2, "EIP network", "Limit STP to drives",
        "Beta / ERD NC limits",
        "X TBIO INPUT1/2; Z TBIO INPUT3/4; Theta X31.5/6 optional",
        "22-24 AWG STP", "—", "No", "—",
        1, "As required", "—", "24 V sinking NC",
        "Shield at drive end only. Limits are drive-side — not ESP32 GPIO.",
        "docs/EXPECTED_ELECTROMECHANICAL_ASSEMBLY.md",
    ),
]

ORDER_PROVIDED_ROWS = [
    ("McMc", "P0078269", "L2", "2198-TBIO", 1, "X axis", "Kinetix 5100 I/O terminal block", "Accepts 16-30 AWG field wiring for limits / field I/O", "Field-wire NC limits -> TBIO INPUT1/2. Motion via EIP - no PTI harness."),
    ("McMc", "P0078269", "L3", "2198-K51CK-D15M", 1, "X axis", "15-pin feedback connector kit", "Terminates 2090-CFBM7DF at E1020 MFB port", "Kit only — pairs with L7 feedback cable."),
    ("McMc", "P0078269", "L6", "2090-CPWM7DF-16AF03", 1, "X axis", "Motor power cable", "16 AWG continuous-flex, 3 m, MPL-A320P", "Drive ↔ motor U/V/W + PE."),
    ("McMc", "P0078269", "L7", "2090-CFBM7DF-CEAF03", 1, "X axis", "Motor feedback cable", "3 m serial feedback, MPL-A320P", "Use with L3 connector kit at drive end."),
    ("McMc", "P0078269", "L11", "2198-K51CK-D15M", 1, "Z axis", "15-pin feedback connector kit", "Terminates feedback at E1004 MFB port", "Kit only — pairs with L15 feedback cable."),
    ("McMc", "P0078269", "L12", "2198-TBIO", 1, "Z axis", "Kinetix 5100 I/O terminal block", "Same as X TBIO", "Field-wire NC limits -> TBIO INPUT3/4. Motion via EIP - no PTI harness."),
    ("McMc", "P0078269", "L14", "2090-CPWM7DF-16AF03", 1, "Z axis", "Motor power cable", "16 AWG continuous-flex, 3 m, MPL-A310F", "Drive ↔ motor U/V/W + PE."),
    ("McMc", "P0078269", "L15", "2090-CFBM7DF-CEAF03", 1, "Z axis", "Motor feedback cable", "3 m serial feedback, MPL-A310F", "Use with L11 connector kit at drive end."),
    ("Youngblood", "P077662", "349-104", "KA GLT1706-LK-00500-1", 1, "Theta", "ERD power cable", "5 m SCHUNK factory cable", "ERD ↔ HCS01 motor power."),
    ("Youngblood", "P077662", "349-544", "KA WWN1208-GK-00500-K02", 1, "Theta", "ERD encoder cable", "5 m HIPERFACE cable", "ERD ↔ HCS01 feedback."),
]

SHOPPING_ROWS = [
    ("WIZ850io / W5500 module", "1", "—", "WIZ850io", "Controller box", "EIP originator SPI Ethernet — see BOM §9.0"),
    ("Cat5e patch (EIP daisy-chain)", "2–3", "Cat5e+", "Generic", "W5500↔X↔Z↔PC", "Exclusive uplink with PC tools"),
    ("Limit switch STP to TBIO", "1 kit", "22 AWG STP", "Belden 8723", "Actuator → TBIO", "NC sinking; INPUT1-4 — not ESP32"),
    ("Motor power cable", "0", "16 AWG continuous-flex", "2090-CPWM7DF-16AF03", "3 m each", "On order ×2 — do not use loose 16 AWG for motor runs"),
    ("Motor feedback cable", "0", "Factory feedback pairs", "2090-CFBM7DF-CEAF03", "3 m each", "On order ×2"),
    ("Feedback connector kit", "0", "—", "2198-K51CK-D15M", "—", "On order ×2"),
    ("Kinetix TBIO", "0", "16-30 AWG terminals", "2198-TBIO", "—", "On order ×2"),
    ("ERD power + encoder cable", "0", "SCHUNK factory", "KA GLT1706 / KA WWN1208", "5 m each", "On order — do not substitute"),
    ("Panel breakers 1489-M2D series", "0", "—", "On-hand (also McMc L8/L9/L17/L18)", "4 pcs", "See 'Panel breakers' sheet — one of each rating"),
    ("AC branch X (40 A)", "1", "8 AWG + ground", "—", "Panel to E1020", "Wire NOT in inventory — breaker M2D400 is on hand"),
    ("AC branch Z (10 A)", "0", "16 AWG + ground", "On-hand 16 AWG", "Panel to E1004", "Breaker M2D100 on hand"),
    ("HCS01 mains branch", "0", "16 AWG + ground", "On-hand 16 AWG", "Panel to HCS01", "Breaker M2D020 on hand — verify peak vs 2 A"),
    ("24 V PSU AC feed", "0", "16 AWG + ground", "On-hand 16 AWG", "Panel to 1606-XL", "Use >=2 A breaker (M2D010 undersized)"),
    ("24 VDC supply", "1", "—", "1606-XLxxx (typical)", "Panel", "PSU not in inventory — purchase"),
    ("24 V to Kinetix CP (×2 drives)", "0", "16 or 18 AWG", "On-hand", "~2 m each", "Use 18 AWG preferred; 16 AWG OK"),
    ("Gripper solenoid valve wire", "0", "18 AWG", "On-hand 18 AWG", "As required", "GPIO4 → CR1 → valve"),
    ("2-pair 22 AWG (limits / SPI)", "~4–6", "2-pair 22 AWG", "On-hand", "Various", "Drive limits + short W5500 SPI"),
    ("Pneumatic tube 6 mm", "~3 m", "6 mm OD", "—", "Gripper", "Not wire — purchase if not using M3 plate bores"),
]


def build_shopping(ws) -> None:
    ws.title = "Still need to source"
    headers = [
        "Item", "Qty still needed", "Wire / cable size", "Source / part no.",
        "Typical length", "Notes",
    ]
    widths = [28, 14, 22, 34, 16, 48]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"
    for i, row in enumerate(SHOPPING_ROWS, start=2):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = WRAP
        qty = str(row[1])
        note = str(row[5]).lower()
        if qty == "0":
            ws.cell(row=i, column=2).fill = PROVIDED_FILL
        elif qty == "1" and "8 awg" in str(row[2]).lower():
            ws.cell(row=i, column=2).fill = NOT_PROVIDED_FILL
        elif "on-hand" in str(row[3]).lower() or "inventory" in note:
            for col in (2, 3, 4):
                ws.cell(row=i, column=col).fill = INVENTORY_OK_FILL


INVENTORY_MAP_ROWS = [
    (
        "1489-M2D400", "X axis AC — 2198-E1020-ERS", "Panel bus → line filter → E1020",
        "Yes — assign this breaker", "40 A matches Rockwell E1020 recommendation. Pair with 8 AWG branch wire (buy). McMc L9.",
    ),
    (
        "1489-M2D100", "Z axis AC — 2198-E1004-ERS", "Panel bus → line filter → E1004",
        "Yes — assign this breaker", "10 A on McMc order (Rockwell table lists 15 A as alternate). Pair with 16 AWG on hand. McMc L18.",
    ),
    (
        "1489-M2D020", "Theta AC — HCS01.1E-W0005", "Panel bus → HCS01 L/N",
        "Yes — assign this breaker", "~2 A cont / 5 A peak drive. 2 A breaker may trip on hard accel — monitor at commissioning. McMc L8.",
    ),
    (
        "1489-M2D010", "Control power AC feed", "Panel bus → 24 VDC PSU (1606-XL) AC input",
        "Yes — assign this breaker", "1 A AC branch for DIN 24 V supply (~120 W class). DC side fuses protect Kinetix CP + valve. McMc L17.",
    ),
    (
        "16 AWG", "Motor power X/Z", "2090-CPWM7DF on order",
        "No — use factory cable", "Ordered 16 AWG flex has drive/motor connectors; loose wire is not a substitute.",
    ),
    (
        "16 AWG", "AC branch — X axis (40 A breaker)", "Panel → 2198-E1020-ERS",
        "No — buy 8 AWG", "40 A branch requires ~8 AWG Cu. 16 AWG is not code-safe at 40 A.",
    ),
    (
        "16 AWG", "AC branch — Z axis (10 A breaker)", "Panel → 2198-E1004-ERS",
        "Yes", "10 A circuit: 16 AWG ampacity is adequate. TBIO/drive terminals accept down to 12 AWG.",
    ),
    (
        "16 AWG", "AC branch — HCS01 Theta", "Panel → HCS01",
        "Yes", "~2 A continuous / 5 A peak. 16 AWG is fine for branch ≤15 A.",
    ),
    (
        "16 AWG", "24 VDC to Kinetix CP (+/−) ×2", "24 V PSU → E1020 & E1004",
        "Yes (18 AWG preferred)", "Both gauges meet 16–24 AWG spec. 18 AWG is easier at TBIO-sized terminals.",
    ),
    (
        "18 AWG", "W5500 SPI harness (short)", "WT32 ↔ WIZ850io",
        "Yes", "MOSI12/MISO35/SCLK5/CS15/RST14. Keep <0.5 m.",
    ),
    (
        "18 AWG", "Limit switches → drive TBIO", "Beta actuators → INPUT1-4",
        "Yes", "NC sinking to drives; not ESP32. 22 AWG STP preferred.",
    ),
    (
        "18 AWG", "Gripper 24 V solenoid valve", "GPIO4 → CR1 → valve coil",
        "Yes", "Within 18–22 AWG spec. Size fuse to coil nameplate.",
    ),
    (
        "Cat5e+", "EIP daisy-chain", "W5500 → X → Z → PC",
        "Yes", "Class 1 EtherNet/IP. Exclusive with PC prove tools.",
    ),
    (
        "2-pair 22 AWG", "X limits MIN/MAX", "Beta 100-ZRS → TBIO INPUT1/2",
        "Yes — 1 cable", "Forward/Reverse Limit in KNX5100C.",
    ),
    (
        "2-pair 22 AWG", "Z limits MIN/MAX", "Beta 80-SRS → TBIO INPUT3/4",
        "Yes — 1 cable", "Forward/Reverse Limit in KNX5100C.",
    ),
    (
        "2-pair 22 AWG", "W5500 SPI (1 cable)", "WT32 ↔ WIZ850io",
        "Yes — 1 cable", "4–5 wires short run. No MCP production path.",
    ),
]

INVENTORY_SUMMARY_ROWS = [
    ("1489-M2D400", "Assign to", "X Kinetix E1020 AC input (via 2198-DB127-F line filter)"),
    ("1489-M2D100", "Assign to", "Z Kinetix E1004 AC input (via 2198-DB111-F line filter)"),
    ("1489-M2D020", "Assign to", "HCS01 Theta drive AC input"),
    ("1489-M2D010", "Assign to", "AC feed to 24 VDC control PSU (not DC load)"),
    ("16 AWG on hand", "Use for", "Z + Theta AC branches, 24 V PSU AC feed, optional 24 V DC runs"),
    ("16 AWG on hand", "Do NOT use for", "X 40 A AC branch (needs 8 AWG); motor U/V/W (use ordered 2090 cables)"),
    ("18 AWG on hand", "Use for", "All TBIO I/O, limits, gripper valve, 24 V DC distribution from PSU"),
    ("2-pair 22 AWG on hand", "Minimum count", "~10 cables (encoders deferred)"),
    ("2-pair 22 AWG on hand", "With encoders (future)", "~12 cables total"),
    ("Still must buy", "—", "8 AWG (+ ground) for X AC branch; 24 VDC PSU; W5500 if missing; Cat5e EIP jumpers; limit STP; pneumatic hose if needed"),
]

PANEL_BREAKER_ROWS = [
    (
        "1489-M2D400", "40 A", "X axis", "2198-E1020-ERS",
        "2198-DB127-F → drive L1/L2/L3", "8 AWG Cu + ground",
        "On hand + McMc L9", "Highest load axis (MPL-A320P). Branch wire is the only missing item.",
    ),
    (
        "1489-M2D100", "10 A", "Z axis", "2198-E1004-ERS",
        "2198-DB111-F → drive L1/L2/L3", "16 AWG Cu + ground",
        "On hand + McMc L18", "Rockwell fuse table also lists 15 A — 10 A is what was ordered.",
    ),
    (
        "1489-M2D020", "2 A", "Theta", "HCS01.1E-W0005-A-03",
        "Panel → HCS01 mains", "16 AWG Cu + ground",
        "On hand + McMc L8", "Drive 2 A cont / 5 A peak. If nuisance trips, review Rexroth startup profile.",
    ),
    (
        "1489-M2D010", "1 A", "Control power", "24 VDC PSU (1606-XLxxx)",
        "Panel → PSU AC input", "16 AWG Cu + ground",
        "On hand + McMc L17", "Protects AC side of 24 V supply. Size PSU ≥3 A DC for two Kinetix CP circuits + valve.",
    ),
]


def build_inventory_map(ws) -> None:
    ws.title = "On-hand inventory map"
    ws["A1"] = "Mapping on-hand wire to gantry runs"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = "Inventory: 1489-M2D010/020/100/400 | 16 AWG | 18 AWG | 2-pair 22 AWG"
    ws["A2"].font = SECTION_FONT
    ws.merge_cells("A2:E2")

    headers = ["On-hand wire", "Application", "Route", "Use inventory?", "Notes"]
    widths = [14, 32, 36, 16, 48]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    for i, row in enumerate(INVENTORY_MAP_ROWS, start=5):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = WRAP
            if col == 4:
                text = str(val)
                if text.startswith("Yes"):
                    cell.fill = INVENTORY_OK_FILL
                elif text.startswith("No"):
                    cell.fill = NOT_PROVIDED_FILL
                else:
                    cell.fill = PARTIAL_FILL

    summary_start = 5 + len(INVENTORY_MAP_ROWS) + 2
    ws.cell(row=summary_start, column=1, value="Quick summary").font = SECTION_FONT
    for j, row in enumerate(INVENTORY_SUMMARY_ROWS, start=summary_start + 1):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=j, column=col, value=val)
            cell.alignment = WRAP
            if col == 1 and "must buy" in str(val).lower():
                cell.fill = NOT_PROVIDED_FILL
            elif col == 3 and "Do NOT" in str(row[1]):
                cell.fill = NOT_PROVIDED_FILL
            elif col == 3 and j > summary_start:
                cell.fill = INVENTORY_OK_FILL


def build_panel_breakers(ws) -> None:
    ws.title = "Panel breakers"
    ws["A1"] = "1489-M2D miniature circuit breaker assignment"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:H1")
    ws["A2"] = (
        "All four ratings on hand (same parts also on McMc order L8/L9/L17/L18). "
        "1489-M2D is AC branch protection — 24 V to drives is fed from a separate DIN PSU."
    )
    ws["A2"].alignment = WRAP
    ws.merge_cells("A2:H2")

    headers = [
        "Part number", "Rating", "Subsystem", "Load",
        "Route", "Branch wire", "Stock / order", "Notes",
    ]
    widths = [16, 8, 12, 28, 28, 18, 18, 44]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A5"

    for i, row in enumerate(PANEL_BREAKER_ROWS, start=5):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = WRAP
            cell.fill = INVENTORY_OK_FILL

    note = 5 + len(PANEL_BREAKER_ROWS) + 1
    ws.cell(row=note, column=1, value="24 VDC distribution (downstream of PSU, not through 1489-M2D):").font = SECTION_FONT
    dc_notes = [
        "Kinetix E1020 + E1004 control power (CP+/CP−): ~1.4 A each @ 24 V — use 18 AWG from PSU; total PSU ≥3 A DC recommended.",
        "Gripper solenoid valve: 18 AWG from same 24 V bus; fuse or PTC per coil current.",
        "WT32-ETH01: separate 5 V/USB or 3.3 V logic supply — not on these AC breakers.",
    ]
    for j, text in enumerate(dc_notes, start=note + 1):
        ws.cell(row=j, column=1, value=text).alignment = WRAP
        ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=8)


def build_order_provided(ws) -> None:
    ws.title = "Order-provided cabling"
    ws["A1"] = "Cables, connector kits, and I/O blocks already on purchase orders"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:J1")
    headers = [
        "Vendor", "Customer PO", "Line", "Part number", "Qty", "Axis",
        "Description", "Specification", "Notes",
    ]
    widths = [12, 14, 8, 32, 6, 10, 28, 36, 40]
    for col, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A4"
    for i, row in enumerate(ORDER_PROVIDED_ROWS, start=4):
        for col, val in enumerate(row, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            cell.alignment = WRAP
            cell.fill = PROVIDED_FILL
    note_row = 4 + len(ORDER_PROVIDED_ROWS) + 1
    ws.cell(row=note_row, column=1, value="Not included on either order:").font = SECTION_FONT
    not_included = [
        "Panel AC branch wire (breakers on hand — see 'Panel breakers' sheet; X branch still needs 8 AWG)",
        "24 VDC PSU and DC wiring to Kinetix CP terminals",
        "WIZ850io / W5500 module and short SPI harness (if not already installed)",
        "Cat5e EIP daisy-chain jumpers (W5500 → X → Z → PC)",
        "Limit switch wiring from Beta actuators → TBIO INPUT1-4 (NC sinking)",
        "Gripper pneumatic hose (6 mm) and 24 V solenoid valve + GPIO4→CR1 wiring",
        "Qualified safety relay / STO design (see docs/HV_LV_SCHEMATICS.md review register)",
    ]
    for j, text in enumerate(not_included, start=note_row + 1):
        ws.cell(row=j, column=1, value=text).alignment = WRAP
        ws.merge_cells(start_row=j, start_column=1, end_row=j, end_column=9)
        ws.cell(row=j, column=1).fill = NOT_PROVIDED_FILL


def main() -> None:
    wb = Workbook()
    build_summary(wb.active)
    build_order_provided(wb.create_sheet())
    build_inventory_map(wb.create_sheet())
    build_panel_breakers(wb.create_sheet())
    build_detail_sheet(wb.create_sheet(), "X axis", X_ROWS)
    build_detail_sheet(wb.create_sheet(), "Z axis", Z_ROWS)
    build_detail_sheet(wb.create_sheet(), "Theta", THETA_ROWS)
    build_detail_sheet(wb.create_sheet(), "EIP network", IF_BOARD_ROWS)
    build_detail_sheet(wb.create_sheet(), "Gripper", GRIPPER_ROWS)
    build_detail_sheet(wb.create_sheet(), "Controller", CONTROLLER_ROWS)
    build_shopping(wb.create_sheet())
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
